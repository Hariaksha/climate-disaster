"""
Phase 2 -- Post-hoc audit sample (paper/attribution_codebook.md, Step 5).

Draws a ~100-article, independent, stratified sample from the now-fully
LLM-coded Master sheet for a final human spot-check, and writes it to a new
`Audit Sample` sheet in attribution.xlsx.

This is NOT the calibration sample (the 50 articles used to clear the
reliability gate before scaling up) -- those are explicitly excluded, since
the audit needs to be independent of the sample used to develop/tune the
codebook. The reliability statistics computed from THIS sample (weighted
and unweighted Cohen's kappa, mean absolute score difference) are the ones
that get reported in the paper's methods section.

Sampling design (documented here, not just in code, so the choice is
auditable):

  1. Build the eligible pool: all Master rows with Archive Status == "found"
     AND not one of the 50 Calibration-sample URLs.
  2. Category E ("contested") is rare in the corpus (10 articles total out
     of 1,591 coded) -- a pure random draw of 100 would very likely contain
     zero or one of them, which can't support a meaningful reliability
     statistic for that category. So all 10 are included automatically.
  3. Category D is also fairly rare (53 articles) -- 20 are randomly
     sampled (not all 53, to leave room for the rest of the design) and
     guaranteed included.
  4. The remaining slots (TARGET_TOTAL - 10 - 20 = 70 by default) are filled
     with a stratified random sample of the rest of the pool (everything
     NOT already pulled into the E/D buckets above -- i.e. categories A/B/C,
     plus Borderline and Trash rows, which never get a category at all).
     Stratified by (Disaster Type, Era), where Era is "pre-2010" / "post-2010"
     per the codebook's calibration-sample stratification approach, with
     slots allocated proportionally to each stratum's share of that pool.

Random seed is fixed (RANDOM_SEED below) so the draw is reproducible --
anyone re-running this script against the same Master data gets the same
sample, which matters for defending the sample wasn't cherry-picked.

The new sheet keeps the LLM's codes visible (Relevance, Attribution
Category, Score, D-Subtype, Contested, Rationale) alongside blank Human-*
columns for you to fill in by hand. Code blind -- i.e. don't look at the
LLM columns while assigning your own codes -- then use a separate script to
compute agreement once you're done. The LLM columns are there for the
*comparison step*, not to anchor your judgment while coding.

Usage:
    python code/sample_audit.py
"""

import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import copy

XLSX_PATH = "attribution.xlsx"
RANDOM_SEED = 42
TARGET_TOTAL = 100
N_CATEGORY_D_SAMPLE = 20  # category E is small enough to take all of it

# Master column indices (must match code/llm_coder.py's layout)
COL_DISASTER_ROW = 1
COL_ARTICLE_NUM = 2
COL_NAME = 3
COL_DISASTER_TYPE = 4
COL_BEGIN_DATE = 5
COL_END_DATE = 6
COL_STATES = 10
COL_URL = 12
COL_TITLE = 13
COL_RELEVANCE = 15
COL_ATTRIBUTION_SCORE = 16
COL_ATTRIBUTION_CATEGORY = 17
COL_D_SUBTYPE = 18
COL_CONTESTED = 19
COL_LLM_RATIONALE = 20
COL_FULL_TEXT = 23
COL_ARCHIVE_STATUS = 25

AUDIT_SHEET_NAME = "Audit Sample"

AUDIT_HEADERS = [
    "Disaster Row #", "Article #", "Name", "Disaster Type", "Begin Date",
    "End Date", "States", "URL", "Title", "Full Article Text",
    "LLM Relevance", "LLM Attribution Category", "LLM Attribution Score",
    "LLM D-Subtype", "LLM Contested", "LLM Rationale",
    "Human Relevance", "Human Attribution Category", "Human Attribution Score",
    "Human D-Subtype", "Human Contested", "Notes",
]


def era_of(begin_date):
    year = int(str(begin_date)[:4])
    return "pre-2010" if year < 2010 else "post-2010"


def main():
    random.seed(RANDOM_SEED)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Master"]
    cal = wb["Calibration"]

    calibration_urls = set()
    for r in range(2, cal.max_row + 1):
        u = cal.cell(row=r, column=8).value
        if u:
            calibration_urls.add(u.strip())

    # 1. Eligible pool: found, coded, not in calibration sample
    pool = []
    for r in range(2, ws.max_row + 1):
        url = ws.cell(row=r, column=COL_URL).value
        status = ws.cell(row=r, column=COL_ARCHIVE_STATUS).value
        relevance = ws.cell(row=r, column=COL_RELEVANCE).value
        if status != "found" or not relevance:
            continue
        if url and url.strip() in calibration_urls:
            continue
        pool.append(r)

    print(f"Eligible pool (found, coded, not calibration sample): {len(pool)} rows")

    category_of = {r: ws.cell(row=r, column=COL_ATTRIBUTION_CATEGORY).value for r in pool}

    e_rows = [r for r in pool if category_of[r] == "E"]
    d_rows = [r for r in pool if category_of[r] == "D"]

    print(f"Category E rows available: {len(e_rows)} (taking all)")
    selected_e = list(e_rows)

    d_sample_n = min(N_CATEGORY_D_SAMPLE, len(d_rows))
    selected_d = random.sample(d_rows, d_sample_n)
    print(f"Category D rows available: {len(d_rows)} (sampling {d_sample_n})")

    already_selected = set(selected_e) | set(selected_d)
    remaining_pool = [r for r in pool if r not in already_selected
                       and category_of[r] not in ("D", "E")]

    remaining_slots = TARGET_TOTAL - len(selected_e) - len(selected_d)
    print(f"Remaining pool for stratified draw: {len(remaining_pool)} rows, "
          f"filling {remaining_slots} slots")

    # 2. Stratify remaining pool by (Disaster Type, Era)
    strata = {}
    for r in remaining_pool:
        dtype = ws.cell(row=r, column=COL_DISASTER_TYPE).value or "Unknown"
        begin = ws.cell(row=r, column=COL_BEGIN_DATE).value
        era = era_of(begin) if begin else "Unknown"
        strata.setdefault((dtype, era), []).append(r)

    total_remaining = len(remaining_pool)
    selected_stratified = []
    # Allocate proportionally, then randomly sample within each stratum
    allocations = {}
    for key, rows in strata.items():
        share = len(rows) / total_remaining if total_remaining else 0
        allocations[key] = round(share * remaining_slots)

    # Adjust rounding so totals land close to remaining_slots
    diff = remaining_slots - sum(allocations.values())
    keys_sorted = sorted(strata.keys(), key=lambda k: -len(strata[k]))
    i = 0
    while diff != 0 and keys_sorted:
        key = keys_sorted[i % len(keys_sorted)]
        if diff > 0:
            allocations[key] += 1
            diff -= 1
        elif allocations[key] > 0:
            allocations[key] -= 1
            diff += 1
        i += 1

    for key, n in allocations.items():
        rows = strata[key]
        n = min(n, len(rows))
        selected_stratified.extend(random.sample(rows, n))

    print(f"Stratified draw selected: {len(selected_stratified)} rows across "
          f"{len(strata)} (Disaster Type, Era) strata")

    final_selection = sorted(set(selected_e) | set(selected_d) | set(selected_stratified))
    print(f"Final audit sample size: {len(final_selection)}")

    # 3. Write to a new sheet
    if AUDIT_SHEET_NAME in wb.sheetnames:
        del wb[AUDIT_SHEET_NAME]
    audit_ws = wb.create_sheet(AUDIT_SHEET_NAME)

    header_ref = ws.cell(row=1, column=1)
    header_font = copy.copy(header_ref.font)
    header_fill = copy.copy(header_ref.fill)
    header_alignment = copy.copy(header_ref.alignment)
    header_border = copy.copy(header_ref.border)

    for c, header in enumerate(AUDIT_HEADERS, 1):
        cell = audit_ws.cell(row=1, column=c)
        cell.value = header
        cell.font = copy.copy(header_font)
        cell.fill = copy.copy(header_fill)
        cell.alignment = copy.copy(header_alignment)
        cell.border = copy.copy(header_border)

    for out_r, master_r in enumerate(final_selection, 2):
        values = [
            ws.cell(row=master_r, column=COL_DISASTER_ROW).value,
            ws.cell(row=master_r, column=COL_ARTICLE_NUM).value,
            ws.cell(row=master_r, column=COL_NAME).value,
            ws.cell(row=master_r, column=COL_DISASTER_TYPE).value,
            ws.cell(row=master_r, column=COL_BEGIN_DATE).value,
            ws.cell(row=master_r, column=COL_END_DATE).value,
            ws.cell(row=master_r, column=COL_STATES).value,
            ws.cell(row=master_r, column=COL_URL).value,
            ws.cell(row=master_r, column=COL_TITLE).value,
            ws.cell(row=master_r, column=COL_FULL_TEXT).value,
            ws.cell(row=master_r, column=COL_RELEVANCE).value,
            ws.cell(row=master_r, column=COL_ATTRIBUTION_CATEGORY).value,
            ws.cell(row=master_r, column=COL_ATTRIBUTION_SCORE).value,
            ws.cell(row=master_r, column=COL_D_SUBTYPE).value,
            ws.cell(row=master_r, column=COL_CONTESTED).value,
            ws.cell(row=master_r, column=COL_LLM_RATIONALE).value,
            None, None, None, None, None, None,  # blank Human-* + Notes columns
        ]
        for c, v in enumerate(values, 1):
            audit_ws.cell(row=out_r, column=c).value = v

    audit_ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)
    print(f"Wrote {len(final_selection)} rows to '{AUDIT_SHEET_NAME}' sheet in {XLSX_PATH}.")


if __name__ == "__main__":
    main()
