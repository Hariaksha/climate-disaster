"""
Phase 2 — Full-sample LLM coding of Master sheet articles.

Calls Claude Sonnet 4.6 on every row in the `Master` sheet of attribution.xlsx
that:
  - has Archive Status == "found" (i.e. has Full Article Text), AND
  - is not already one of the 50 rows hand-coded in the `Calibration` sheet
    (matched by URL, since calibration already cleared the reliability gate
    -- see paper/attribution_codebook.md, Calibration & Validation Procedure),
    AND
  - does not already have a Relevance code in Master (so the script is
    resume-safe: re-running it only codes rows that are still blank).

For each selected row it asks the model to apply the two-step procedure from
paper/attribution_codebook.md:
  1. Relevance screening -> Relevant / Borderline / Trash
  2. Attribution coding (Relevant only) -> category A-E, plus the auxiliary
     D-subtype and Contested (E-flag) tags

Results are written back into the Master sheet (column layout as of the
current Master header row -- Attribution Category/D-Subtype/Contested/LLM
Rationale sit right next to Relevance/Attribution Score, pushing National
Newspaper/Calibration Sample/Full Article Text/Archive URL/Archive
Status/Word Count five slots to the right of where they used to be):
  - col O (15) Relevance              -- existing column
  - col P (16) Attribution Score       -- existing column (header relabeled
                                          from "Attribution Index" to match
                                          codebook/Calibration terminology --
                                          "Attribution Index" is reserved for
                                          the disaster-level aggregate in the
                                          `All Disasters` sheet). Holds the
                                          NUMERIC score (-1, 0, 0.5, 1),
                                          because the existing `All
                                          Disasters` AVERAGEIFS formula over
                                          Master!P already assumes this
                                          column is numeric. Written as a
                                          literal value, not an Excel
                                          formula -- the Calibration sheet's
                                          formula-based score column
                                          (=IF(category="A",1,...)) only
                                          shows a value once Excel has
                                          actually recalculated it, which
                                          makes it unreadable to any script
                                          that opens the workbook with
                                          openpyxl(data_only=True) before
                                          that happens (confirmed first-hand
                                          on this workbook -- see kappa
                                          calculation note). A literal value
                                          computed once, from the same code
                                          path that writes the category, is
                                          always in sync and always
                                          readable. Left blank for
                                          Borderline/Trash, matching the
                                          codebook ("only Relevant articles
                                          receive an attribution code").
  - col Q (17) Attribution Category   -- NEW. The letter A-E itself -- the
                                          primary/source value; Attribution
                                          Score above is mechanically derived
                                          from it in the same write step, so
                                          the mapping is auditable and so
                                          Cohen's kappa can be computed
                                          against human codes later.
  - col R (18) D-Subtype              -- NEW. "natural_variability" or
                                          "other_human_causes", category D only.
  - col S (19) Contested              -- NEW. TRUE for category E, else FALSE.
  - col T (20) LLM Rationale          -- NEW. 1-2 sentence justification, for
                                          spot-checking/audit, not analysis.

SECURITY NOTE: This script reads the API key from the ANTHROPIC_API_KEY
environment variable. It deliberately does NOT read it from CLAUDE.md or
any other file in this repo. CLAUDE.md is currently untracked (not
committed) in this repo, which avoids the worst risk, but it's still a
live secret sitting in a plain file -- one stray `git add .` away from
being committed and pushed to github.com/Hariaksha/climate-disaster.
Prefer setting ANTHROPIC_API_KEY as a shell/environment variable, or in a
local .env file added to .gitignore, instead of leaving it in CLAUDE.md.

Usage:
    pip install anthropic openpyxl
    export ANTHROPIC_API_KEY=sk-ant-...   # set in your shell, not in a file
    python code/llm_coder.py              # TEST_MODE first, then full run

Cost tracking: the script sums actual input/output tokens from each API
response and prints a running cost estimate against MODEL_PRICING below,
and will refuse to start a new call once MAX_BUDGET_USD is exceeded.
"""

import os
import sys
import time
import json
import random

import openpyxl

try:
    import anthropic
except ImportError:
    sys.exit(
        "Missing dependency. Run: pip install anthropic openpyxl"
    )

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

XLSX_PATH = "attribution.xlsx"
MODEL = "claude-sonnet-4-6"

# Pricing as of this writing (verify at https://platform.claude.com/docs/en/about-claude/pricing
# before a large run -- prices can change).
PRICE_INPUT_PER_MTOK = 3.00
PRICE_OUTPUT_PER_MTOK = 15.00

MAX_BUDGET_USD = 30.00       # hard stop
CHECKPOINT_EVERY = 25        # save workbook every N coded rows
TEST_MODE = False            # True -> only code TEST_MODE_ROWS rows, then stop
TEST_MODE_ROWS = 5

MAX_RETRIES = 5
BASE_BACKOFF_SECONDS = 2.0

# Master sheet column indices (1-based, matches attribution.xlsx header row).
# NOTE: these were reordered in the spreadsheet so the attribution columns
# sit next to Relevance/Attribution Score for readability -- if you reorder
# columns again, update this block to match (or this script will write into
# the wrong columns).
COL_DISASTER_ROW = 1
COL_ARTICLE_NUM = 2
COL_NAME = 3
COL_DISASTER_TYPE = 4
COL_BEGIN_DATE = 5
COL_END_DATE = 6
COL_STATES = 10
COL_URL = 12
COL_TITLE = 13
COL_RELEVANCE = 15
COL_ATTRIBUTION_SCORE = 16
COL_ATTRIBUTION_CATEGORY = 17
COL_D_SUBTYPE = 18
COL_CONTESTED = 19
COL_LLM_RATIONALE = 20
COL_NATIONAL_NEWSPAPER = 21
COL_CALIBRATION_SAMPLE = 22
COL_FULL_TEXT = 23
COL_ARCHIVE_URL = 24
COL_ARCHIVE_STATUS = 25
COL_WORD_COUNT = 26

NEW_COLUMN_HEADERS = {
    COL_ATTRIBUTION_CATEGORY: "Attribution Category",
    COL_D_SUBTYPE: "D-Subtype",
    COL_CONTESTED: "Contested",
    COL_LLM_RATIONALE: "LLM Rationale",
}

CATEGORY_TO_SCORE = {"A": 1.0, "B": 0.5, "C": 0.0, "D": -1.0, "E": 0.0}

# ---------------------------------------------------------------------------
# Prompt, built from paper/attribution_codebook.md
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """You are coding newspaper articles for the research project \
"Reading the Storm: Media Attribution of U.S. Climate Disasters, 2000-2024." \
You will be given metadata about a specific NOAA-documented billion-dollar \
disaster event, and the full text of one article retrieved as candidate \
coverage of that event. Apply the two-step procedure below exactly.

STEP 1 - RELEVANCE SCREENING
Decide whether the article is actually about THIS SPECIFIC disaster event \
(correct disaster type, correct geography/state, and a date consistent with \
the event window, including reasonable lead-up or retrospective coverage).

- Relevant: the article substantively discusses this specific event -- its \
impacts, recovery, costs, forecasts, or causal framing.
- Borderline: plausibly about this event but ambiguous -- e.g. a wire brief \
too thin to confirm the event match, a "wildfire season" or "hurricane \
season" roundup that may or may not cover this specific event, or a \
retrospective piece (1+ years later) that mentions the event only in passing.
- Trash: off-topic, wrong event/year/disaster type, a paywall/login/cookie \
page, an aggregator or archive listing page with no real article content, a \
dead link, or a photo gallery with only captions (captions cannot sustain \
causal framing).

STEP 2 - ATTRIBUTION CODING (Relevant articles only)
If, and only if, Step 1 = Relevant, assign exactly one attribution category:

- A (score +1.0) Explicit anthropogenic attribution: the article directly \
and affirmatively links the event's occurrence, frequency, or intensity to \
human-caused climate change / global warming. Often cites climate \
scientists, formal attribution studies (e.g. World Weather Attribution), or \
uses confident causal language. Example: "Climate scientists say the \
unprecedented rainfall was supercharged by a warmer atmosphere, which holds \
more moisture."
- B (score +0.5) Hedged / contextual climate mention: climate change is \
raised as a possible or partial factor, with hedged language ("may," \
"could," "some scientists believe"), or mentioned briefly/contextually \
without being central to the causal framing. Example: "The fire comes amid \
a years-long drought that some link to a changing climate."
- C (score 0.0) No causal attribution: pure event reporting -- impacts, \
damages, casualties, response, recovery, forecasts -- with no discussion of \
broader cause (climate or otherwise). Example: "The storm knocked out power \
to 500,000 homes across the Southeast."
- D (score -1.0) Explicit non-climate attribution: the article attributes \
the event to natural variability/cyclical weather patterns (El Nino/La \
Nina, "typical for this time of year," historical cycles) OR to non-climate \
human factors (floodplain development, forest management, infrastructure \
failures) -- and in doing so explicitly downplays or rejects a climate-\
change link. Example: "Meteorologists say this is a fairly typical La Nina \
winter pattern, not evidence of a longer-term trend." For category D only, \
also set d_subtype to "natural_variability" or "other_human_causes" \
depending on which of the two it is.
- E (score 0.0, contested) Mixed / contested: the article presents both \
climate-linked and non-climate framings with roughly comparable weight \
(e.g. a climate scientist and a skeptical official both quoted). Example: \
"While some researchers point to climate change as a contributing factor, \
local officials say the flooding was mainly due to outdated drainage \
infrastructure." Set contested = true only for category E.

If Step 1 is Borderline or Trash, attribution_category, d_subtype, and \
contested must all be null/false -- do not guess a category for an article \
you did not mark Relevant.

Always call the `code_article` tool with your answer. Keep rationale to 1-2 \
sentences -- it is for audit/spot-checking, not for the dataset itself."""

CODE_ARTICLE_TOOL = {
    "name": "code_article",
    "description": "Record the relevance and attribution coding for one article.",
    "input_schema": {
        "type": "object",
        "properties": {
            "relevance": {
                "type": "string",
                "enum": ["Relevant", "Borderline", "Trash"],
            },
            "attribution_category": {
                "type": ["string", "null"],
                "enum": ["A", "B", "C", "D", "E", None],
                "description": "Required if relevance == Relevant, else null.",
            },
            "d_subtype": {
                "type": ["string", "null"],
                "enum": ["natural_variability", "other_human_causes", None],
                "description": "Set only if attribution_category == D, else null.",
            },
            "contested": {
                "type": "boolean",
                "description": "True only if attribution_category == E.",
            },
            "rationale": {
                "type": "string",
                "description": "1-2 sentence justification for audit purposes.",
            },
        },
        "required": ["relevance", "contested", "rationale"],
    },
}


def build_user_message(ws, row):
    name = ws.cell(row=row, column=COL_NAME).value
    disaster_type = ws.cell(row=row, column=COL_DISASTER_TYPE).value
    begin_date = ws.cell(row=row, column=COL_BEGIN_DATE).value
    end_date = ws.cell(row=row, column=COL_END_DATE).value
    states = ws.cell(row=row, column=COL_STATES).value
    title = ws.cell(row=row, column=COL_TITLE).value
    url = ws.cell(row=row, column=COL_URL).value
    full_text = ws.cell(row=row, column=COL_FULL_TEXT).value or ""

    return f"""DISASTER EVENT METADATA
Name: {name}
Disaster type: {disaster_type}
Event window: {begin_date} to {end_date}
States affected: {states}

ARTICLE
URL: {url}
Title: {title}

Full article text:
\"\"\"
{full_text}
\"\"\"

Code this article per the procedure in your instructions."""


# ---------------------------------------------------------------------------
# Cost tracking
# ---------------------------------------------------------------------------

class Budget:
    def __init__(self, max_usd):
        self.max_usd = max_usd
        self.spent_usd = 0.0
        self.calls = 0

    def add(self, input_tokens, output_tokens):
        cost = (input_tokens / 1e6) * PRICE_INPUT_PER_MTOK + (
            output_tokens / 1e6
        ) * PRICE_OUTPUT_PER_MTOK
        self.spent_usd += cost
        self.calls += 1
        return cost

    def exceeded(self):
        return self.spent_usd >= self.max_usd


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def call_with_retry(client, **kwargs):
    for attempt in range(MAX_RETRIES):
        try:
            return client.messages.create(**kwargs)
        except anthropic.RateLimitError:
            wait = BASE_BACKOFF_SECONDS * (2 ** attempt) + random.uniform(0, 1)
            print(f"  rate limited, retrying in {wait:.1f}s...")
            time.sleep(wait)
        except anthropic.APIStatusError as e:
            if e.status_code >= 500 and attempt < MAX_RETRIES - 1:
                wait = BASE_BACKOFF_SECONDS * (2 ** attempt) + random.uniform(0, 1)
                print(f"  server error {e.status_code}, retrying in {wait:.1f}s...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError("Exceeded max retries")


def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit(
            "ANTHROPIC_API_KEY is not set. Set it in your shell "
            "(export ANTHROPIC_API_KEY=sk-ant-...) -- do not hardcode it "
            "here or put it in a file that gets committed to git."
        )

    client = anthropic.Anthropic(api_key=api_key)

    print(f"Loading {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Master"]
    cal = wb["Calibration"]

    # Ensure new column headers exist
    for col, header in NEW_COLUMN_HEADERS.items():
        if ws.cell(row=1, column=col).value != header:
            ws.cell(row=1, column=col).value = header

    # Relabel the existing numeric-score column to match the codebook's/
    # Calibration sheet's terminology -- it holds the article-level
    # Attribution Score, not the disaster-level Attribution Index (that
    # name is reserved for the aggregated AttributionIndex_d in the
    # `All Disasters` sheet). Purely cosmetic: the All Disasters formula
    # references this column by letter (Master!P), not by header text, so
    # renaming the header does not affect it.
    ws.cell(row=1, column=COL_ATTRIBUTION_SCORE).value = "Attribution Score"

    # Rows already hand-coded in Calibration, matched by URL -- skip these,
    # the calibration gate already covers them.
    calibration_urls = set()
    for r in range(2, cal.max_row + 1):
        u = cal.cell(row=r, column=8).value  # Calibration col H = URL
        if u:
            calibration_urls.add(u.strip())

    # Select candidate rows
    candidates = []
    for r in range(2, ws.max_row + 1):
        url = ws.cell(row=r, column=COL_URL).value
        status = ws.cell(row=r, column=COL_ARCHIVE_STATUS).value
        full_text = ws.cell(row=r, column=COL_FULL_TEXT).value
        already_coded = ws.cell(row=r, column=COL_RELEVANCE).value
        if not url or status != "found" or not full_text or already_coded:
            continue
        if url.strip() in calibration_urls:
            continue
        candidates.append(r)

    print(f"{len(candidates)} rows queued for coding "
          f"(found, not in Calibration sample, not yet coded).")

    if TEST_MODE:
        candidates = candidates[:TEST_MODE_ROWS]
        print(f"TEST_MODE on -- limiting to first {len(candidates)} rows. "
              f"Set TEST_MODE = False for the full run.")

    budget = Budget(MAX_BUDGET_USD)
    coded_since_save = 0

    for i, row in enumerate(candidates, 1):
        if budget.exceeded():
            print(f"Budget cap (${MAX_BUDGET_USD:.2f}) reached after "
                  f"{budget.calls} calls / ${budget.spent_usd:.2f} spent. Stopping.")
            break

        user_msg = build_user_message(ws, row)

        response = call_with_retry(
            client,
            model=MODEL,
            max_tokens=400,
            system=SYSTEM_PROMPT,
            tools=[CODE_ARTICLE_TOOL],
            tool_choice={"type": "tool", "name": "code_article"},
            messages=[{"role": "user", "content": user_msg}],
        )

        cost = budget.add(response.usage.input_tokens, response.usage.output_tokens)

        tool_use_block = next(
            (b for b in response.content if b.type == "tool_use"), None
        )
        if tool_use_block is None:
            print(f"  row {row}: no tool_use in response, skipping")
            continue

        result = tool_use_block.input
        relevance = result.get("relevance")
        category = result.get("attribution_category")
        d_subtype = result.get("d_subtype")
        contested = result.get("contested", False)
        rationale = result.get("rationale", "")

        # Defensive normalization: enforce the codebook's constraints
        # ourselves rather than trusting the model never deviates from the
        # system prompt's instructions (it's enum-constrained, but cheap
        # insurance against e.g. a Borderline row coming back with a
        # leftover d_subtype/contested value).
        if relevance != "Relevant":
            category = None
            d_subtype = None
            contested = False
        if category != "D":
            d_subtype = None
        if category != "E":
            contested = False

        ws.cell(row=row, column=COL_RELEVANCE).value = relevance
        if relevance == "Relevant" and category in CATEGORY_TO_SCORE:
            ws.cell(row=row, column=COL_ATTRIBUTION_SCORE).value = CATEGORY_TO_SCORE[category]
            ws.cell(row=row, column=COL_ATTRIBUTION_CATEGORY).value = category
        else:
            ws.cell(row=row, column=COL_ATTRIBUTION_SCORE).value = None
            ws.cell(row=row, column=COL_ATTRIBUTION_CATEGORY).value = None
        ws.cell(row=row, column=COL_D_SUBTYPE).value = d_subtype
        ws.cell(row=row, column=COL_CONTESTED).value = bool(contested)
        ws.cell(row=row, column=COL_LLM_RATIONALE).value = rationale

        coded_since_save += 1
        print(f"[{i}/{len(candidates)}] row {row}: {relevance}"
              f"{'/' + category if category else ''}"
              f"  (+${cost:.4f}, running total ${budget.spent_usd:.2f})")

        if coded_since_save >= CHECKPOINT_EVERY:
            wb.save(XLSX_PATH)
            print(f"  checkpoint saved ({budget.calls} calls so far).")
            coded_since_save = 0

    wb.save(XLSX_PATH)
    print(f"Done. {budget.calls} calls made, ${budget.spent_usd:.2f} spent.")


if __name__ == "__main__":
    main()
