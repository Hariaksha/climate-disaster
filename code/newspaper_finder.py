"""
Newspaper Website Finder — Large Dataset Edition (3000+ rows)
- Uses ddgs library (renamed from duckduckgo-search)
- Filters out Wikipedia, social media, and directory results
- Skips already-processed rows (resume-safe)
- Exponential backoff on rate limits
- Randomized delays to avoid bot detection

Install dependencies:
    pip install requests openpyxl ddgs
"""

import time
import random
import requests
from openpyxl import load_workbook
from ddgs import DDGS
from ddgs.exceptions import RatelimitException

# ── CONFIG ───────────────────────────────────────────────────────────────────
SPREADSHEET_PATH = "/Users/hariaksha/Documents/GitHub/climate-disaster/Access World News Database.xlsx"  # <-- change to your filename
SHEET_NAME       = None               # None = active sheet; or e.g. "Sheet1"
REQUEST_TIMEOUT  = 8
MIN_DELAY        = 2.0
MAX_DELAY        = 4.5
SAVE_EVERY       = 5
MAX_RETRIES      = 4
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}

# Domains to reject — not real newspaper websites
BLOCKED_DOMAINS = [
    "wikipedia.org",
    "wikidata.org",
    "facebook.com",
    "twitter.com",
    "x.com",
    "instagram.com",
    "linkedin.com",
    "youtube.com",
    "yelp.com",
    "newspapers.com",    # archive/subscription aggregator
    "genealogybank.com",
    "ancestry.com",
    "findmypast.com",
    "newsbank.com",
    "proquest.com",
    "loc.gov",           # Library of Congress
    "chroniclingamerica",
]


def is_blocked(url: str) -> bool:
    """Return True if the URL belongs to a domain we should skip."""
    url_lower = url.lower()
    return any(blocked in url_lower for blocked in BLOCKED_DOMAINS)


def search_for_url(title: str, city: str, state: str, language: str) -> str | None:
    """
    Search DDG for the newspaper's official website.
    Appends -wikipedia to the query and skips blocked domains.
    Returns first acceptable result URL, or None.
    """
    parts = [f'"{title}"']          # quote the title for exact matching
    if city and city.strip() not in ("-", ""):
        parts.append(city)
    if state:
        parts.append(state)
    parts.append("newspaper")
    parts.append("-wikipedia")        # exclude Wikipedia results
    parts.append("-site:wikipedia.org")
    if language and language.lower() not in ("english", ""):
        parts.append(language)

    query = " ".join(parts)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            with DDGS() as ddgs:
                # Fetch a few results so we can skip blocked ones
                results = list(ddgs.text(query, max_results=5))

            for r in results:
                url = r.get("href", "")
                if url and not is_blocked(url):
                    return url

            # All top results were blocked — return None
            return None

        except RatelimitException:
            wait = (2 ** attempt) + random.uniform(5, 15)
            print(f"    [rate limited] waiting {wait:.0f}s before retry {attempt}/{MAX_RETRIES}...")
            time.sleep(wait)

        except Exception as e:
            print(f"    [search error] {e}")
            return None

    print(f"    [gave up after {MAX_RETRIES} retries]")
    return None


def verify_url(url: str) -> tuple[str, str]:
    """Fetch the URL to confirm it is live. Returns (final_url, status_label)."""
    if not url:
        return ("", "not found")
    if not url.startswith("http"):
        url = "https://" + url

    parked_keywords = [
        "godaddy", "sedo.com", "dan.com", "afternic",
        "this domain", "domain for sale", "buy this domain",
        "parked by", "namecheap parking",
    ]

    try:
        resp = requests.get(
            url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True
        )
        final_url = resp.url
        body_lower = resp.text[:3000].lower()

        if any(kw in body_lower for kw in parked_keywords):
            return (final_url, "parked/unavailable")
        if resp.status_code == 200:
            if final_url.rstrip("/") != url.rstrip("/"):
                return (final_url, "verified (redirected)")
            return (url, "verified")
        elif resp.status_code == 404:
            return (url, "not found")
        else:
            return (url, f"error {resp.status_code}")

    except requests.exceptions.SSLError:
        try:
            http_url = url.replace("https://", "http://")
            resp = requests.get(
                http_url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True
            )
            if resp.status_code == 200:
                return (http_url, "verified (http only)")
        except Exception:
            pass
        return (url, "SSL error")
    except requests.exceptions.ConnectionError:
        return (url, "not found")
    except requests.exceptions.Timeout:
        return (url, "timeout")
    except Exception as e:
        return (url, f"error: {str(e)[:50]}")


def main():
    wb = load_workbook(SPREADSHEET_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    if ws["E1"].value != "Website URL":
        ws["E1"] = "Website URL"
    if ws["F1"].value != "Status":
        ws["F1"] = "Status"

    total_rows = ws.max_row
    skipped = 0
    processed = 0

    print(f"Spreadsheet has {total_rows - 1} data rows.")
    print("Rows with an existing status in column F will be skipped (resume-safe).")
    print("-" * 60)

    for row_idx in range(2, total_rows + 1):
        title           = str(ws.cell(row_idx, 1).value or "").strip()
        city            = str(ws.cell(row_idx, 2).value or "").strip()
        state           = str(ws.cell(row_idx, 3).value or "").strip()
        language        = str(ws.cell(row_idx, 4).value or "").strip()
        existing_status = ws.cell(row_idx, 6).value

        # Resume: skip rows already processed
        if existing_status and str(existing_status).strip() not in ("", "None"):
            skipped += 1
            continue

        if not title or title == "None":
            ws.cell(row_idx, 5).value = ""
            ws.cell(row_idx, 6).value = "skipped (no title)"
            continue

        progress = f"[{row_idx-1}/{total_rows-1}]"
        print(f"{progress} {title} | {city}, {state}", end=" ... ", flush=True)

        found_url = search_for_url(title, city, state, language)
        url, status = verify_url(found_url)

        ws.cell(row_idx, 5).value = url
        ws.cell(row_idx, 6).value = status
        processed += 1

        print(f"{status} → {url[:70] if url else '—'}")

        if processed % SAVE_EVERY == 0:
            wb.save(SPREADSHEET_PATH)
            print(f"  [progress saved — {processed} processed, {skipped} already done]")

        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    wb.save(SPREADSHEET_PATH)
    print("\n" + "=" * 60)
    print(f"Done! Processed: {processed} | Skipped (already done): {skipped}")
    print(f"Spreadsheet saved to: {SPREADSHEET_PATH}")


if __name__ == "__main__":
    main()
