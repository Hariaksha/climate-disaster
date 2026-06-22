"""
Post-hoc audit reliability statistics (paper/attribution_codebook.md, Step 5).

Computes the independent human-vs-LLM reliability statistics for the
100-row Audit Sample (see code/sample_audit.py), to be reported in the
paper's Methods section in place of the `[TBD]` marker:

  1. Relevance (3-category: Relevant/Borderline/Trash): unweighted Cohen's
     kappa, on all 100 audited rows.
  2. Attribution (5-category: A/B/C/D/E): unweighted Cohen's kappa, plus a
     weighted kappa, on the subset of rows where BOTH the human coder and
     the LLM independently classified the article as Relevant (only there
     do both raters have a meaningful 5-category attribution code).
  3. Mean absolute difference (MAD) between human- and LLM-assigned
     numeric attribution scores, on that same Relevant-by-both subset.

Weighted kappa here is computed from the *actual numeric attribution
scores* underlying each category (A=+1, B=+0.5, C=0, D=-1, E=0), not from
naive category rank. This matters because C and E both score 0 -- a
C-vs-E disagreement should count as a zero-weight (no real) disagreement
in the weighted scheme, exactly as the codebook intends when it says
unweighted kappa "penalizes a B-vs-C disagreement the same as an A-vs-D
disagreement" and weighted kappa should fix that. Both linear (weight
proportional to |score difference|) and quadratic (proportional to the
squared difference) weighting are reported.

Usage:
    python code/audit_reliability.py
"""

import numpy as np
import openpyxl

XLSX_PATH = "attribution.xlsx"
AUDIT_SHEET = "Audit Sample"

COL_HUMAN_RELEVANCE = 17
COL_HUMAN_CATEGORY = 18
COL_HUMAN_SCORE = 19
COL_LLM_RELEVANCE = 11
COL_LLM_CATEGORY = 12
COL_LLM_SCORE = 13

SCORE_MAP = {"A": 1.0, "B": 0.5, "C": 0.0, "D": -1.0, "E": 0.0}


def unweighted_kappa(rater1, rater2, categories):
    """Standard (unweighted) Cohen's kappa for categorical labels."""
    idx = {c: i for i, c in enumerate(categories)}
    k = len(categories)
    n = len(rater1)
    O = np.zeros((k, k))
    for a, b in zip(rater1, rater2):
        O[idx[a], idx[b]] += 1
    O /= n
    p_row, p_col = O.sum(axis=1), O.sum(axis=0)
    po = np.trace(O)
    pe = float(np.sum(p_row * p_col))
    return 1.0 if pe == 1 else (po - pe) / (1 - pe)


def score_weighted_kappa(cat1, cat2, categories, score_map, power):
    """Weighted kappa (Cohen 1968) using disagreement weights based on the
    actual numeric attribution score difference between categories (raised
    to `power`: 1 = linear, 2 = quadratic), not naive category rank.
    """
    idx = {c: i for i, c in enumerate(categories)}
    k = len(categories)
    n = len(cat1)
    scores = np.array([score_map[c] for c in categories])
    O = np.zeros((k, k))
    for a, b in zip(cat1, cat2):
        O[idx[a], idx[b]] += 1
    O /= n
    p_row, p_col = O.sum(axis=1), O.sum(axis=0)
    E = np.outer(p_row, p_col)
    max_diff = scores.max() - scores.min()
    W = np.zeros((k, k))
    for i in range(k):
        for j in range(k):
            d = abs(scores[i] - scores[j])
            W[i, j] = (d / max_diff) ** power if max_diff > 0 else 0.0
    num = float(np.sum(W * O))
    den = float(np.sum(W * E))
    return 1.0 if den == 0 else 1 - num / den


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[AUDIT_SHEET]

    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({
            "human_rel": ws.cell(row=r, column=COL_HUMAN_RELEVANCE).value,
            "human_cat": ws.cell(row=r, column=COL_HUMAN_CATEGORY).value,
            "human_score": ws.cell(row=r, column=COL_HUMAN_SCORE).value,
            "llm_rel": ws.cell(row=r, column=COL_LLM_RELEVANCE).value,
            "llm_cat": ws.cell(row=r, column=COL_LLM_CATEGORY).value,
            "llm_score": ws.cell(row=r, column=COL_LLM_SCORE).value,
        })
    n_total = len(rows)
    print(f"Audit sample: N = {n_total} rows\n")

    # ---- 1. Relevance: unweighted kappa, all rows ----
    rel_categories = ["Relevant", "Borderline", "Trash"]
    human_rel = [x["human_rel"] for x in rows]
    llm_rel = [x["llm_rel"] for x in rows]
    assert all(v in rel_categories for v in human_rel + llm_rel), "Unexpected Relevance value"

    kappa_rel = unweighted_kappa(human_rel, llm_rel, rel_categories)
    agree_rel = sum(1 for h, l in zip(human_rel, llm_rel) if h == l) / n_total
    print("=== Relevance (3-category) ===")
    print(f"Raw agreement: {agree_rel:.1%}")
    print(f"Unweighted Cohen's kappa: {kappa_rel:.4f}")
    print(f"Target (codebook): kappa >= 0.70 (\"substantial\")\n")

    # ---- 2. Attribution: subset where BOTH raters said Relevant ----
    both_relevant = [x for x in rows if x["human_rel"] == "Relevant" and x["llm_rel"] == "Relevant"]
    n_attr = len(both_relevant)
    cat_categories = ["A", "B", "C", "D", "E"]
    human_cat = [x["human_cat"] for x in both_relevant]
    llm_cat = [x["llm_cat"] for x in both_relevant]
    assert all(v in cat_categories for v in human_cat + llm_cat), "Unexpected Attribution category value"

    kappa_attr_unw = unweighted_kappa(human_cat, llm_cat, cat_categories)
    kappa_attr_lin = score_weighted_kappa(human_cat, llm_cat, cat_categories, SCORE_MAP, power=1)
    kappa_attr_quad = score_weighted_kappa(human_cat, llm_cat, cat_categories, SCORE_MAP, power=2)
    agree_attr = sum(1 for h, l in zip(human_cat, llm_cat) if h == l) / n_attr

    human_scores = np.array([SCORE_MAP[c] for c in human_cat])
    llm_scores = np.array([SCORE_MAP[c] for c in llm_cat])
    mad = float(np.mean(np.abs(human_scores - llm_scores)))

    print("=== Attribution (5-category, A-E) ===")
    print(f"N (Relevant by both human and LLM): {n_attr} of {n_total}")
    print(f"Raw agreement: {agree_attr:.1%}")
    print(f"Unweighted Cohen's kappa: {kappa_attr_unw:.4f}")
    print(f"Linear score-weighted kappa: {kappa_attr_lin:.4f}")
    print(f"Quadratic score-weighted kappa: {kappa_attr_quad:.4f}")
    print(f"Mean absolute difference (human vs LLM score, [-1,1] scale): {mad:.4f}")
    print(f"Targets (codebook): unweighted/weighted kappa >= 0.70, ideally weighted >= 0.80\n")

    # ---- Confusion matrix for transparency ----
    print("=== Attribution confusion matrix (rows=Human, cols=LLM) ===")
    print("      " + "  ".join(f"{c:>3s}" for c in cat_categories))
    for hc in cat_categories:
        row_counts = [sum(1 for h, l in zip(human_cat, llm_cat) if h == hc and l == lc) for lc in cat_categories]
        print(f"  {hc} | " + "  ".join(f"{v:3d}" for v in row_counts))

    print("\n=== Relevance confusion matrix (rows=Human, cols=LLM) ===")
    print("      " + "  ".join(f"{c:>10s}" for c in rel_categories))
    for hc in rel_categories:
        row_counts = [sum(1 for h, l in zip(human_rel, llm_rel) if h == hc and l == lc) for lc in rel_categories]
        print(f"  {hc:10s} | " + "  ".join(f"{v:10d}" for v in row_counts))


if __name__ == "__main__":
    main()
