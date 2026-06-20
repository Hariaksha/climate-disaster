#!/usr/bin/env python3
"""
Phase 1 (v4 — Wayback Machine + stub detection)

Archives all Master sheet URLs via the Wayback Machine and extracts the full
article text directly, verifying it's a real article and not a paywall stub.

Replaces the archive.ph approach, which requires solving a CAPTCHA on every
single request and cannot be automated.

For each URL:
  1. Check the Wayback Machine availability API for an existing snapshot.
  2. If none exists, submit a Save Page Now request (free, no CAPTCHA).
  3. Fetch the snapshot HTML and extract the main article text (trafilatura).
  4. If extracted text has >= MIN_WORDS words, mark 'found' — ready for
     Phase 2 LLM coding, with the full text written directly to col 19.
  5. If too short (paywall stub) or no snapshot exists at all, mark
     'needs_factiva' — you'll need to pull these manually.

Results written to the Master sheet:
  Col 19: Full Article Text   — extracted text (only for 'found' rows)
  Col 20: Archive URL         — Wayback Machine snapshot URL
  Col 21: Archive Status      — 'found', 'needs_factiva', or 'failed'
  Col 22: Word Count          — word count of extracted text (for QA)

Fully resumable — rows already marked 'found' or 'needs_factiva' are skipped.
"""

import time
import re
import requests
import trafilatura
import openpyxl
from datetime import datetime

# ── Config ────────────────────────────────────────────────────────────────────
WORKBOOK_PATH = '/Users/hariaksha/Documents/GitHub/climate-disaster/attribution.xlsx'
SHEET_NAME    = 'Master'
URL_COL          = 12  # Column L: original article URL
TEXT_COL         = 19  # Column S: Full Article Text
ARCHIVE_URL_COL  = 20  # Column T: Wayback Machine snapshot URL
ARCHIVE_STATUS_COL = 21  # Column U: 'found' / 'needs_factiva' / 'failed'
WORD_COUNT_COL   = 22  # Column V: extracted word count

MIN_WORDS = 150   # below this, treat as a paywall stub / teaser, not a real article
MIN_WORDS_NPR_TRANSCRIPT = 30  # NPR's official transcripts are authoritative even if short
NPR_ID_PATTERN = re.compile(r'npr\.org/(?:[^/]+/){0,4}?(\d{5,})(?:/|$)')
CNN_FILEID_PATTERN = re.compile(r'transcripts\.cnn\.com/show/([a-z]+)\??.*start_fileid=([a-z]+)_(\d{4}-\d{2}-\d{2})_(\d+)')

DELAY_CHECK  = 1.5   # seconds between availability checks
DELAY_FETCH  = 1.0   # seconds between snapshot fetches
DELAY_SAVE_WAIT = 8  # seconds to wait after submitting Save Page Now
MAX_RETRIES  = 2

CONSEC_FAIL_THRESHOLD = 8
BACKOFF_SCHEDULE = [60, 180, 300]   # 1, 3, 5 minutes — WBM is much gentler than archive.ph
BACKOFF_COUNTDOWN_INTERVAL = 30

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/120.0.0.0 Safari/537.36'
    ),
}

WAYBACK_AVAIL_API = 'https://archive.org/wayback/available'
WAYBACK_SAVE_URL  = 'https://web.archive.org/save/'
# ─────────────────────────────────────────────────────────────────────────────

consecutive_failures = 0
backoff_level = 0


def countdown_pause(seconds: int, reason: str):
    print(f'\n⏸  {reason}')
    print(f'   Pausing for {seconds // 60}m {seconds % 60}s — script will resume automatically.')
    end_time = time.time() + seconds
    while True:
        remaining = int(end_time - time.time())
        if remaining <= 0:
            break
        mins, secs = divmod(remaining, 60)
        print(f'   Resuming in {mins}m {secs:02d}s ...', end='\r', flush=True)
        time.sleep(min(BACKOFF_COUNTDOWN_INTERVAL, remaining))
    print(f'\n▶  Resuming now.\n')


def handle_rate_limit():
    global backoff_level, consecutive_failures
    pause_secs = BACKOFF_SCHEDULE[min(backoff_level, len(BACKOFF_SCHEDULE) - 1)]
    backoff_level = min(backoff_level + 1, len(BACKOFF_SCHEDULE) - 1)
    consecutive_failures = 0
    countdown_pause(pause_secs, f'Rate limit detected — auto-pausing for {pause_secs // 60} minutes.')


def is_rate_limited(response) -> bool:
    if response is None:
        return False
    return response.status_code in (429, 503)


def check_existing(url: str, session: requests.Session, _retried=False):
    """Check Wayback Machine availability API. Returns (snapshot_url_or_None, rate_limited)."""
    try:
        r = session.get(WAYBACK_AVAIL_API, params={'url': url}, headers=HEADERS, timeout=20)
        if is_rate_limited(r):
            return None, True
        if r.status_code == 200:
            data = r.json()
            snap = data.get('archived_snapshots', {}).get('closest', {})
            if snap.get('available') and snap.get('url'):
                return snap['url'], False
    except Exception:
        if not _retried:
            time.sleep(3)
            return check_existing(url, session, _retried=True)
    return None, False


def submit_new(url: str, session: requests.Session):
    """Submit Save Page Now request. Returns (snapshot_url_or_None, rate_limited)."""
    try:
        r = session.get(f'{WAYBACK_SAVE_URL}{url}', headers=HEADERS, timeout=90, allow_redirects=True)
        if is_rate_limited(r):
            return None, True
        if 'web.archive.org/web/' in r.url:
            return r.url, False
    except requests.exceptions.Timeout:
        pass
    except Exception:
        pass

    # Save didn't redirect immediately (server error, still processing, etc.)
    # Wait and re-check via the availability API before giving up — the save
    # may have completed in the background, or an existing snapshot may have
    # been missed due to a transient error.
    time.sleep(12)
    return check_existing(url, session)


def fetch_and_extract(snapshot_url: str, session: requests.Session):
    """Fetch the snapshot HTML and extract main article text. Returns (text, word_count)."""
    try:
        r = session.get(snapshot_url, headers=HEADERS, timeout=30)
        if r.status_code != 200:
            return '', 0
        text = trafilatura.extract(
            r.text, url=snapshot_url, include_comments=False, include_tables=False
        )
        if not text:
            return '', 0
        return text, len(text.split())
    except Exception:
        return '', 0


def try_npr_transcript(original_url: str, session: requests.Session):
    """
    For npr.org URLs where the article page is a short teaser, NPR often has
    the full official transcript at npr.org/transcripts/{id}. Try that as a
    fallback. Returns (snapshot_url, text, word_count) or (None, '', 0).
    """
    if 'npr.org' not in original_url:
        return None, '', 0
    m = NPR_ID_PATTERN.search(original_url)
    if not m:
        return None, '', 0
    npr_id = m.group(1)
    transcript_url = f'https://www.npr.org/transcripts/{npr_id}'
    if transcript_url.rstrip('/') == original_url.rstrip('/'):
        return None, '', 0  # already tried this exact URL

    snap_url, rate_limited = check_existing(transcript_url, session)
    if rate_limited:
        handle_rate_limit()
        snap_url, rate_limited = check_existing(transcript_url, session)
    if not snap_url:
        snap_url, rate_limited = submit_new(transcript_url, session)
        if rate_limited:
            handle_rate_limit()
            snap_url, rate_limited = submit_new(transcript_url, session)

    if snap_url:
        text, wc = fetch_and_extract(snap_url, session)
        if wc >= MIN_WORDS_NPR_TRANSCRIPT:
            return snap_url, text, wc
    return None, '', 0


def try_cnn_modern_url(original_url: str, session: requests.Session):
    """
    transcripts.cnn.com has two URL formats: a dead legacy query-string format
    (?start_fileid=show_date_NN) that falls back to a generic index page, and
    a working format (/show/{show}/date/{date}/segment/{NN}) with the same
    show/date/segment info. Reconstruct the working URL from the dead one.
    Returns (snapshot_url, text, word_count) or (None, '', 0).
    """
    m = CNN_FILEID_PATTERN.search(original_url)
    if not m:
        return None, '', 0
    show, _show2, date, seg = m.groups()
    new_url = f'https://transcripts.cnn.com/show/{show}/date/{date}/segment/{int(seg):02d}'

    snap_url, rate_limited = check_existing(new_url, session)
    if rate_limited:
        handle_rate_limit()
        snap_url, rate_limited = check_existing(new_url, session)
    if not snap_url:
        snap_url, rate_limited = submit_new(new_url, session)
        if rate_limited:
            handle_rate_limit()
            snap_url, rate_limited = submit_new(new_url, session)

    if snap_url:
        text, wc = fetch_and_extract(snap_url, session)
        if wc >= MIN_WORDS:
            return snap_url, text, wc
    return None, '', 0


def process_url(url: str, session: requests.Session) -> tuple:
    """
    Returns (snapshot_url, status, text, word_count).
    status is 'found', 'needs_factiva', or 'failed'.
    """
    global consecutive_failures, backoff_level
    rl_used = 0

    # Step 1: check existing snapshot
    snap_url, rate_limited = check_existing(url, session)
    if rate_limited:
        rl_used += 1
        if rl_used <= 1:
            handle_rate_limit()
            snap_url, rate_limited = check_existing(url, session)

    if snap_url:
        text, wc = fetch_and_extract(snap_url, session)
        if wc >= MIN_WORDS:
            consecutive_failures = 0
            backoff_level = max(0, backoff_level - 1)
            return snap_url, 'found', text, wc
        else:
            # Found a snapshot but it's too short — likely a stub. Try saving a fresh copy.
            print(f'      → Snapshot found but only {wc} words (stub?) — trying fresh save...')

    # Step 2: submit Save Page Now (covers both "not found" and "stub" cases)
    print(f'      → Submitting Save Page Now...')
    new_snap_url, rate_limited = submit_new(url, session)
    if rate_limited:
        rl_used += 1
        if rl_used <= 1:
            handle_rate_limit()
            new_snap_url, rate_limited = submit_new(url, session)

    if new_snap_url:
        time.sleep(DELAY_SAVE_WAIT)
        text, wc = fetch_and_extract(new_snap_url, session)
        if wc >= MIN_WORDS:
            consecutive_failures = 0
            backoff_level = max(0, backoff_level - 1)
            return new_snap_url, 'found', text, wc
        else:
            # Still short — try source-specific fallbacks before giving up
            npr_snap, npr_text, npr_wc = try_npr_transcript(url, session)
            if npr_snap:
                consecutive_failures = 0
                return npr_snap, 'found', npr_text, npr_wc
            cnn_snap, cnn_text, cnn_wc = try_cnn_modern_url(url, session)
            if cnn_snap:
                consecutive_failures = 0
                return cnn_snap, 'found', cnn_text, cnn_wc
            return new_snap_url, 'needs_factiva', '', wc

    # Nothing usable found at all — last resort: source-specific fallbacks
    npr_snap, npr_text, npr_wc = try_npr_transcript(url, session)
    if npr_snap:
        consecutive_failures = 0
        return npr_snap, 'found', npr_text, npr_wc
    cnn_snap, cnn_text, cnn_wc = try_cnn_modern_url(url, session)
    if cnn_snap:
        consecutive_failures = 0
        return cnn_snap, 'found', cnn_text, cnn_wc

    if snap_url:
        # We had an old stub snapshot but couldn't save a better one
        return snap_url, 'needs_factiva', '', 0

    return None, 'failed', '', 0


def add_headers_if_missing(ws):
    headers = {
        TEXT_COL: 'Full Article Text',
        ARCHIVE_URL_COL: 'Archive URL',
        ARCHIVE_STATUS_COL: 'Archive Status',
        WORD_COUNT_COL: 'Word Count',
    }
    for col, name in headers.items():
        if ws.cell(row=1, column=col).value != name:
            ws.cell(row=1, column=col).value = name


def main():
    print(f'Loading workbook: {WORKBOOK_PATH}')
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]
    add_headers_if_missing(ws)

    total_rows = ws.max_row - 1
    session = requests.Session()

    global consecutive_failures, backoff_level
    found = 0
    needs_factiva = 0
    failed = 0
    skipped = 0

    start_time = datetime.now()
    print(f'Starting at {start_time.strftime("%H:%M:%S")} — {total_rows} articles to process\n')

    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=URL_COL).value
        existing_status = ws.cell(row=row_idx, column=ARCHIVE_STATUS_COL).value

        article_num = row_idx - 1
        title = ws.cell(row=row_idx, column=13).value or ''

        if not url:
            skipped += 1
            continue

        if existing_status in ('found', 'needs_factiva'):
            skipped += 1
            if article_num % 100 == 0:
                print(f'[{article_num:4d}/{total_rows}] ... skipping already-processed rows ...')
            continue

        print(f'[{article_num:4d}/{total_rows}] Processing | {title[:60]}')
        print(f'      URL: {str(url)[:80]}')

        snap_url, status, text, wc = None, 'failed', '', 0
        for attempt in range(1, MAX_RETRIES + 1):
            if attempt > 1:
                print(f'      Retry {attempt}/{MAX_RETRIES}...')
                time.sleep(5)
            snap_url, status, text, wc = process_url(str(url), session)
            if status in ('found', 'needs_factiva'):
                break

        ws.cell(row=row_idx, column=ARCHIVE_URL_COL).value = snap_url
        ws.cell(row=row_idx, column=ARCHIVE_STATUS_COL).value = status
        ws.cell(row=row_idx, column=WORD_COUNT_COL).value = wc
        if status == 'found':
            ws.cell(row=row_idx, column=TEXT_COL).value = text[:32000]  # Excel cell limit safety
        wb.save(WORKBOOK_PATH)

        if status == 'found':
            found += 1
            print(f'      ✓ Found full text ({wc} words): {snap_url}')
        elif status == 'needs_factiva':
            needs_factiva += 1
            print(f'      ⚠ Stub/paywalled ({wc} words) — needs Factiva: {snap_url}')
        else:
            failed += 1
            consecutive_failures += 1
            print(f'      ✗ No snapshot available — needs Factiva')
            if consecutive_failures >= CONSEC_FAIL_THRESHOLD:
                handle_rate_limit()

        print()
        time.sleep(DELAY_CHECK)

    elapsed = datetime.now() - start_time
    print('=' * 60)
    print(f'DONE in {elapsed}')
    print(f'  Full text found:        {found:4d}')
    print(f'  Needs Factiva (stub):   {needs_factiva:4d}')
    print(f'  Needs Factiva (failed): {failed:4d}')
    print(f'  Skipped:                {skipped:4d}')
    total_needs_factiva = needs_factiva + failed
    print(f'\n  Total needing Factiva: {total_needs_factiva}/{total_rows - skipped}')
    print(f'\nWorkbook saved to: {WORKBOOK_PATH}')
    print('\nNext step: run phase2_llm_coding.py on rows marked "found".')
    print('For rows marked "needs_factiva", pull full text manually and paste into col 19.')


if __name__ == '__main__':
    main()
