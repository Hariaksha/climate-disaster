#!/usr/bin/env python3
"""
Newspaper Website Finder — Perplexity Sonar Edition (Final)
Uses the Perplexity Sonar API (web-grounded LLM) to find each newspaper's
official website, then verifies the URL with an HTTP check.

Cost estimate for 3,648 rows using sonar model:
  ~$19-22 total for full run

Install:
    pip install openai requests openpyxl

Get your API key at: https://www.perplexity.ai/settings/api
"""

import time
import random
import re
import requests
from openai import OpenAI
from openpyxl import load_workbook

# -- CONFIG -------------------------------------------------------------------
SPREADSHEET_PATH   = "/Users/hariaksha/Documents/GitHub/climate-disaster/Access World News Database.xlsx"   # <-- change to your filename
SHEET_NAME         = "Sheet1"                # None = active sheet
PERPLEXITY_API_KEY = "YOUR API KEY HERE" # <-- paste your key here. find in Perplexity API Platform --> Newspaper Finder --> API Keys
MODEL              = "sonar"             # or "sonar-pro" for harder cases
MIN_DELAY          = 1.0                 # seconds between API calls
MAX_DELAY          = 2.0
SAVE_EVERY         = 5
REQUEST_TIMEOUT    = 8                   # seconds for HTTP verification
# -----------------------------------------------------------------------------

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}

client = OpenAI(
    api_key=PERPLEXITY_API_KEY,
    base_url="https://api.perplexity.ai"
)

SYSTEM_PROMPT = (
    "You are a research assistant helping build a database of newspaper and news wire websites. "
    "Given a publication's name, city, state, and language, find the best available URL for it.\n\n"
    "Respond with ONLY one of:\n"
    "  1. A single URL (e.g. https://www.example.com) — the best current web presence for this publication\n"
    "  2. The word NOTFOUND — only if the publication has completely ceased to exist with no web presence\n\n"
    "Guidelines:\n"
    "- Be generous in what you accept as a valid URL. A section page, hub page, or "
    "regional subdirectory of a larger site is acceptable "
    "(e.g. https://apnews.com/hub/alabama for 'Associated Press State Wire: Alabama').\n"
    "- If the publication was absorbed into a larger outlet, return the most specific "
    "URL within that outlet that covers this publication's content.\n"
    "- Use the city and state to disambiguate between publications with the same name "
    "in different locations.\n"
    "- Do NOT return: Wikipedia, Facebook, Twitter, Yelp, newspapers.com, "
    "genealogybank.com, ProQuest, Ancestry, or any archive/aggregator site.\n"
    "- If the domain is parked or for sale, look for an alternative active URL instead.\n"
    "- If the publication is in a language other than English, prioritize its native-language site.\n"
    "- Only return NOTFOUND if you have exhausted all possibilities and the publication "
    "truly has no current web presence whatsoever.\n"
    "- Return ONLY the URL or NOTFOUND. No explanation, no punctuation, no extra text."
)

URL_PATTERN = re.compile(r"https?://[^\s\"'<>]+")


def verify_url(url):
    """Fetch the URL to confirm it is live. Returns (final_url, status_label)."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        if r.status_code == 200:
            return (r.url, "verified")
        elif r.status_code == 404:
            return (url, "not found (model suggested)")
        else:
            return (url, "unverified (" + str(r.status_code) + ")")
    except requests.exceptions.SSLError:
        try:
            http_url = url.replace("https://", "http://")
            r = requests.get(http_url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            if r.status_code == 200:
                return (r.url, "verified (http only)")
        except Exception:
            pass
        return (url, "unverified (SSL error)")
    except requests.exceptions.ConnectionError:
        return (url, "unverified (unreachable)")
    except requests.exceptions.Timeout:
        return (url, "unverified (timeout)")
    except Exception as e:
        return (url, "unverified (" + str(e)[:40] + ")")


def ask_perplexity(title, city, state, language):
    """
    Ask Perplexity Sonar to find the newspaper's official website.
    Returns (url, status).
    """
    city_str = city if city and city not in ("-", "") else "unknown city"
    lang_str = (", " + language) if language and language.lower() != "english" else ""

    user_msg = (
        "Find the official website for this publication:\n"
        "Name: " + title + "\n"
        "Location: " + city_str + ", " + state + lang_str + "\n"
        "Return only the URL or NOTFOUND."
    )

    for attempt in range(2):
        try:
            response = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": user_msg},
                ],
                temperature=0.0,
            )
            answer = response.choices[0].message.content.strip()

            if not answer or answer.upper() == "NOTFOUND":
                return ("", "not found")

            match = URL_PATTERN.search(answer)
            if match:
                url = match.group(0).rstrip(".,;)")
                return verify_url(url)

            return ("", "not found")

        except Exception as e:
            if attempt == 0:
                print("    [API error: " + str(e)[:60] + ", retrying...]")
                time.sleep(5)
            else:
                return ("", "API error: " + str(e)[:40])

    return ("", "not found")


def main():
    if PERPLEXITY_API_KEY == "YOUR_API_KEY_HERE":
        print("ERROR: Please set your PERPLEXITY_API_KEY at the top of the script.")
        print("Get your key at: https://www.perplexity.ai/settings/api")
        return

    wb = load_workbook(SPREADSHEET_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    if ws["E1"].value != "Website URL":
        ws["E1"] = "Website URL"
    if ws["F1"].value != "Status":
        ws["F1"] = "Status"

    total_rows = ws.max_row
    skipped = 0
    processed = 0

    print("Spreadsheet: " + SPREADSHEET_PATH + " | " + str(total_rows - 1) + " data rows")
    print("Model: " + MODEL + " | Estimated cost: ~$19-22 for full run")
    print("Rows with an existing status in column F will be skipped (resume-safe).")
    print("-" * 65)

    for row_idx in range(2, total_rows + 1):

        # -- REMOVE THIS BLOCK AFTER TESTING --
        # if row_idx > 101:
        #     break
        # -- END TEST BLOCK -------------------

        title           = str(ws.cell(row_idx, 1).value or "").strip()
        city            = str(ws.cell(row_idx, 2).value or "").strip()
        state           = str(ws.cell(row_idx, 3).value or "").strip()
        language        = str(ws.cell(row_idx, 4).value or "").strip()
        existing_status = ws.cell(row_idx, 6).value

        # Resume: skip already-processed rows
        if existing_status and str(existing_status).strip() not in ("", "None"):
            skipped += 1
            continue

        if not title or title == "None":
            ws.cell(row_idx, 5).value = ""
            ws.cell(row_idx, 6).value = "skipped (no title)"
            continue

        print("[" + str(row_idx - 1) + "/" + str(total_rows - 1) + "] " + title + " | " + city + ", " + state, end=" ... ", flush=True)

        url, status = ask_perplexity(title, city, state, language)

        ws.cell(row_idx, 5).value = url
        ws.cell(row_idx, 6).value = status
        processed += 1

        print(status + " -> " + (url[:70] if url else "-"))

        if processed % SAVE_EVERY == 0:
            wb.save(SPREADSHEET_PATH)
            print("  [saved -- " + str(processed) + " processed, " + str(skipped) + " skipped]")

        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    wb.save(SPREADSHEET_PATH)
    print("\n" + "=" * 65)
    print("Done! Processed: " + str(processed) + " | Skipped (already done): " + str(skipped))
    print("Spreadsheet saved to: " + SPREADSHEET_PATH)


if __name__ == "__main__":
    main()
    