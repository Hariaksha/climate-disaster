#!/usr/bin/env python3
"""
Phase 1: Archive URL collection for climate disaster articles.

Strategy:
  Step A — CHECK-ONLY pass: GET /newest/[url] for all 1694 articles.
            Fast (4s delay), no rate limiting issues. Finds already-archived articles.
  Step B — SUBMIT pass: POST /submit/ for articles not found in Step A.
            Slower (35s delay), may hit rate limits — runs after Step A is complete.

Run this script once. It will do Step A first (should complete without issues),
then automatically move to Step B for remaining articles.

Results written to attribution.xlsx cols 20 (Archive URL) and 21 (Archive Status).
Status values:
  found      — already on archive.ph, URL recorded
  submitted  — newly submitted and confirmed
  failed     — could not archive (will need manual Factiva retrieval)
"""

import time
import re
import requests
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime

WORKBOOK_PATH       = '/Users/hariaksha/Documents/GitHub/climate-disaster/attribution.xlsx'
SHEET_NAME          = 'Master'
URL_COL             = 12
ARCHIVE_URL_COL     = 20
ARCHIVE_STATUS_COL  = 21

DELAY_CHECK         = 4     # seconds between GET checks (Step A)
DELAY_SUBMIT        = 40    # seconds between POST submissions (Step B)
DELAY_AFTER_SUBMIT  = 40    # seconds to wait before checking if submission landed
CONSEC_FAIL_THRESH  = 5     # consecutive failures before auto-pause
BACKOFF_SCHEDULE    = [5*60, 10*60, 20*60, 30*60]   # escalating pauses in seconds

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/120.0.0.0 Safari/537.36',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

# Global state
consecutive_failures = 0
backoff_level        = 0


# ─── Helpers ────────────────────────────────────────────────────────────────

def countdown_pause(seconds, reason):
    """Display a live countdown while pausing."""
    print(f'\n⏸  {reason} — auto-pausing for {seconds//60}m {seconds%60}s.')
    end = time.time() + seconds
    while True:
        remaining = int(end - time.time())
        if remaining <= 0:
            break
        mins, secs = divmod(remaining, 60)
        print(f'   Resuming in {mins}m {secs:02d}s ...', end='\r')
        time.sleep(min(1, remaining))
    print('\n▶  Resuming now.')


def is_valid_archive_url(url):
    """Return True only if url looks like a real archive.ph snapshot."""
    if not url:
        return False
    url = url.strip()
    if not url.startswith('http'):
        return False
    bad_patterns = ['/submit/', '/search/', '/newest/', 'archive.ph/?', 'archive.ph/wip/']
    return not any(p in url for p in bad_patterns)


def is_rate_limited(response):
    """Detect HTTP 429, 503, Cloudflare blocks, or 'too many requests' body text."""
    if response.status_code in (429, 503):
        return True
    text_lower = response.text.lower()
    if any(p in text_lower for p in ['too many requests', 'rate limit', 'cloudflare', 'access denied']):
        return True
    return False


def handle_rate_limit():
    """Escalating backoff pause on rate limit."""
    global consecutive_failures, backoff_level
    consecutive_failures += 1
    pause = BACKOFF_SCHEDULE[min(backoff_level, len(BACKOFF_SCHEDULE) - 1)]
    backoff_level = min(backoff_level + 1, len(BACKOFF_SCHEDULE) - 1)
    countdown_pause(pause, f'Rate limit detected (failure #{consecutive_failures})')


# ─── Step A: Check-only (GET) ────────────────────────────────────────────────

def check_existing(url, session):
    """
    GET archive.ph/newest/[url].
    Returns (archive_url_or_None, rate_limited_bool).
    """
    check_url = f'https://archive.ph/newest/{url}'
    try:
        r = session.get(check_url, headers=HEADERS, timeout=20, allow_redirects=True)
        if is_rate_limited(r):
            return None, True
        # A successful archive redirect lands on an archive.ph/[hash] URL
        final_url = r.url
        if is_valid_archive_url(final_url) and 'archive.ph' in final_url:
            # Make sure it's not just the /newest/ page itself
            if re.search(r'archive\.ph/[A-Za-z0-9]{4,}', final_url):
                return final_url, False
        return None, False
    except requests.RequestException:
        return None, False


def step_a_check_pass(ws, session):
    """
    Fast check-only pass over all rows.
    Skips rows already marked found/submitted.
    Writes results to cols 20-21 after each hit.
    Returns list of row numbers still needing submission.
    """
    global consecutive_failures, backoff_level

    wb = ws.parent
    total = ws.max_row - 1  # subtract header
    needs_submit = []
    found_count  = 0
    skipped      = 0

    print(f'\n{"="*60}')
    print(f'STEP A — Check-only pass ({total} articles)')
    print(f'{"="*60}\n')

    for row in range(2, ws.max_row + 1):
        url_cell    = ws.cell(row=row, column=URL_COL).value
        status_cell = ws.cell(row=row, column=ARCHIVE_STATUS_COL).value
        title_cell  = ws.cell(row=row, column=13).value or '(no title)'

        if not url_cell:
            continue

        # Already processed — skip
        if status_cell in ('found', 'submitted'):
            skipped += 1
            continue

        idx = row - 1
        print(f'[{idx:4d}/{total}] Checking | {str(title_cell)[:60]}')

        archive_url, rate_limited = check_existing(str(url_cell), session)

        if rate_limited:
            handle_rate_limit()
            # Retry once after pause
            archive_url, rate_limited = check_existing(str(url_cell), session)
            if rate_limited:
                handle_rate_limit()
                archive_url, _ = check_existing(str(url_cell), session)

        if archive_url:
            print(f'        ✓ Found: {archive_url}')
            ws.cell(row=row, column=ARCHIVE_URL_COL).value    = archive_url
            ws.cell(row=row, column=ARCHIVE_STATUS_COL).value = 'found'
            wb.save(WORKBOOK_PATH)
            found_count += 1
            consecutive_failures = 0
            backoff_level = max(0, backoff_level - 1)
        else:
            print(f'        · Not on archive.ph yet')
            needs_submit.append(row)

        time.sleep(DELAY_CHECK)

    print(f'\nStep A complete — {found_count} found, {len(needs_submit)} need submission.')
    return needs_submit


# ─── Step B: Submit new articles ─────────────────────────────────────────────

def submit_new(url, session):
    """
    POST to archive.ph/submit/.
    Returns (archive_url_or_None, rate_limited_bool).
    archive_url is None if submission was accepted but URL not yet known.
    """
    try:
        r = session.post(
            'https://archive.ph/submit/',
            data={'url': url, 'anyway': '1'},
            headers=HEADERS,
            timeout=30,
            allow_redirects=True,
        )
        if is_rate_limited(r):
            return None, True

        # If archive.ph redirected to an existing snapshot, grab it
        if is_valid_archive_url(r.url) and re.search(r'archive\.ph/[A-Za-z0-9]{4,}', r.url):
            return r.url, False

        # Otherwise submission was queued — check for a link in the page
        soup = BeautifulSoup(r.text, 'html.parser')
        for tag in soup.find_all('a', href=True):
            href = tag['href']
            if is_valid_archive_url(href) and re.search(r'archive\.ph/[A-Za-z0-9]{4,}', href):
                return href, False

        # Submission accepted but no URL yet — caller will wait and re-check
        return None, False

    except requests.RequestException:
        return None, False


def step_b_submit_pass(ws, rows_to_submit, session):
    """
    Submit articles not found in Step A.
    Writes results after each article.
    """
    global consecutive_failures, backoff_level

    wb    = ws.parent
    total = len(rows_to_submit)
    done  = 0
    fails = 0

    print(f'\n{"="*60}')
    print(f'STEP B — Submission pass ({total} articles remaining)')
    print(f'{"="*60}\n')

    for i, row in enumerate(rows_to_submit, 1):
        url_cell   = ws.cell(row=row, column=URL_COL).value
        title_cell = ws.cell(row=row, column=13).value or '(no title)'

        if not url_cell:
            continue

        # Re-check: might have been done in a prior interrupted run
        status_cell = ws.cell(row=row, column=ARCHIVE_STATUS_COL).value
        if status_cell in ('found', 'submitted'):
            done += 1
            continue

        url = str(url_cell)
        print(f'[{i:4d}/{total}] Submitting | {str(title_cell)[:55]}')

        # Check once more before submitting (maybe it appeared since Step A)
        archive_url, rate_limited = check_existing(url, session)
        if rate_limited:
            handle_rate_limit()
            archive_url, _ = check_existing(url, session)

        if archive_url:
            print(f'         ✓ Already found: {archive_url}')
            ws.cell(row=row, column=ARCHIVE_URL_COL).value    = archive_url
            ws.cell(row=row, column=ARCHIVE_STATUS_COL).value = 'found'
            wb.save(WORKBOOK_PATH)
            done += 1
            consecutive_failures = 0
            backoff_level = max(0, backoff_level - 1)
            time.sleep(DELAY_CHECK)
            continue

        # Submit
        print(f'         → Posting to archive.ph...')
        archive_url, rate_limited = submit_new(url, session)

        if rate_limited:
            handle_rate_limit()
            # One retry
            archive_url, rate_limited = submit_new(url, session)
            if rate_limited:
                handle_rate_limit()
                archive_url, _ = submit_new(url, session)

        if archive_url:
            print(f'         ✓ Submitted + URL: {archive_url}')
            ws.cell(row=row, column=ARCHIVE_URL_COL).value    = archive_url
            ws.cell(row=row, column=ARCHIVE_STATUS_COL).value = 'submitted'
            wb.save(WORKBOOK_PATH)
            done += 1
            consecutive_failures = 0
            backoff_level = max(0, backoff_level - 1)
            time.sleep(DELAY_SUBMIT)
            continue

        # Submission was queued — wait and check
        print(f'         → Waiting {DELAY_AFTER_SUBMIT}s for archive to process...')
        time.sleep(DELAY_AFTER_SUBMIT)
        archive_url, _ = check_existing(url, session)

        if archive_url:
            print(f'         ✓ Now found: {archive_url}')
            ws.cell(row=row, column=ARCHIVE_URL_COL).value    = archive_url
            ws.cell(row=row, column=ARCHIVE_STATUS_COL).value = 'submitted'
            wb.save(WORKBOOK_PATH)
            done += 1
            consecutive_failures = 0
            backoff_level = max(0, backoff_level - 1)
        else:
            print(f'         ✗ Could not archive — marking failed')
            ws.cell(row=row, column=ARCHIVE_STATUS_COL).value = 'failed'
            wb.save(WORKBOOK_PATH)
            fails += 1
            consecutive_failures += 1
            if consecutive_failures >= CONSEC_FAIL_THRESH:
                countdown_pause(BACKOFF_SCHEDULE[min(backoff_level, len(BACKOFF_SCHEDULE)-1)],
                                f'{CONSEC_FAIL_THRESH} consecutive failures')
                consecutive_failures = 0
                backoff_level = min(backoff_level + 1, len(BACKOFF_SCHEDULE) - 1)

        time.sleep(DELAY_SUBMIT)

    print(f'\nStep B complete — {done} archived, {fails} failed.')
    return fails


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    global consecutive_failures, backoff_level

    print(f'Loading workbook: {WORKBOOK_PATH}')
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    # Ensure output columns have headers
    if not ws.cell(row=1, column=ARCHIVE_URL_COL).value:
        ws.cell(row=1, column=ARCHIVE_URL_COL).value    = 'Archive URL'
        ws.cell(row=1, column=ARCHIVE_STATUS_COL).value = 'Archive Status'
        wb.save(WORKBOOK_PATH)

    start = datetime.now().strftime('%H:%M:%S')
    total = ws.max_row - 1
    print(f'Starting at {start} — {total} articles to process\n')

    session = requests.Session()

    # ── Step A: fast check-only pass ──────────────────────────────────────
    needs_submit = step_a_check_pass(ws, session)

    # Reload workbook in case saves changed it on disk
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    # ── Step B: submit anything not found ─────────────────────────────────
    if needs_submit:
        print(f'\n{len(needs_submit)} articles not found on archive.ph.')
        print('Starting submission pass in 10 seconds (Ctrl+C to stop here)...')
        time.sleep(10)
        step_b_submit_pass(ws, needs_submit, session)
    else:
        print('\nAll articles already on archive.ph — no submissions needed!')

    # ── Final summary ──────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]
    found    = sum(1 for r in range(2, ws.max_row+1)
                   if ws.cell(r, ARCHIVE_STATUS_COL).value == 'found')
    submitted = sum(1 for r in range(2, ws.max_row+1)
                    if ws.cell(r, ARCHIVE_STATUS_COL).value == 'submitted')
    failed   = sum(1 for r in range(2, ws.max_row+1)
                   if ws.cell(r, ARCHIVE_STATUS_COL).value == 'failed')

    print(f'\n{"="*60}')
    print(f'FINAL SUMMARY')
    print(f'{"="*60}')
    print(f'  Already on archive.ph : {found}')
    print(f'  Newly submitted       : {submitted}')
    print(f'  Failed (need Factiva) : {failed}')
    print(f'  Not processed         : {total - found - submitted - failed}')
    print(f'{"="*60}')
    print(f'Done at {datetime.now().strftime("%H:%M:%S")}')


if __name__ == '__main__':
    main()
